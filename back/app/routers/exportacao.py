from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc
import csv
import io
from typing import Optional, Any
from datetime import datetime
from app.core.database import get_db
from app.core.dependencies import get_current_active_user
from app.models.protocolo import Protocolo
from app.models.usuario import Usuario

router = APIRouter()


def _serialize(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M")
    return str(val)


def _build_rows(protocolos: list[Protocolo]) -> tuple[list[str], list[list[str]]]:
    headers = [
        "ID", "Nº Chamado Interno", "Título", "Tipo de Evento", "Tipo de Rede",
        "Tipo de Atendimento", "Status Atendimento", "Estado",
        "POP A", "POP B", "Site A", "Site B",
        "Equipamento A", "Equipamento B", "Porta A", "Porta B",
        "Responsável Trecho", "Responsável Atendimento ID",
        "Nº Chamado OS", "Protocolo Parceiro",
        "Data/Hora Falha", "Data Criação",
        "Clientes Afetados", "Ativo",
    ]
    rows = []
    for p in protocolos:
        rows.append([
            _serialize(p.id),
            _serialize(p.numero_chamado_interno),
            _serialize(p.titulo),
            _serialize(p.tipo_evento),
            _serialize(p.tipo_rede),
            _serialize(p.tipo_atendimento),
            _serialize(p.status_atendimento),
            _serialize(p.estado),
            _serialize(p.nome_pop_a),
            _serialize(p.nome_pop_b),
            _serialize(p.site_a),
            _serialize(p.site_b),
            _serialize(p.equipamento_site_a),
            _serialize(p.equipamento_site_b),
            _serialize(p.porta_site_a),
            _serialize(p.porta_site_b),
            _serialize(p.responsavel_trecho),
            _serialize(p.responsavel_atendimento_id),
            _serialize(p.numero_chamado_os),
            _serialize(p.protocolo_parceiro),
            _serialize(p.data_hora_falha),
            _serialize(p.data_criacao),
            _serialize(p.clientes_afetados),
            "Sim" if p.ativo else "Não",
        ])
    return headers, rows


@router.get("/protocolos/csv")
async def exportar_protocolos_csv(
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
    tipo_evento: Optional[str] = Query(None),
    tipo_rede: Optional[str] = Query(None),
    ativo: Optional[bool] = Query(None),
):
    stmt = select(Protocolo)
    if tipo_evento:
        stmt = stmt.where(Protocolo.tipo_evento == tipo_evento)
    if tipo_rede:
        stmt = stmt.where(Protocolo.tipo_rede == tipo_rede)
    if ativo is not None:
        stmt = stmt.where(Protocolo.ativo == ativo)
    stmt = stmt.order_by(desc(Protocolo.data_criacao))

    result = await db.execute(stmt)
    protocolos = result.scalars().all()

    headers, rows = _build_rows(protocolos)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(headers)
    writer.writerows(rows)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=protocolos.csv",
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


@router.get("/protocolos/excel")
async def exportar_protocolos_excel(
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(get_current_active_user),
    tipo_evento: Optional[str] = Query(None),
    tipo_rede: Optional[str] = Query(None),
    ativo: Optional[bool] = Query(None),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return await exportar_protocolos_csv(db, current_user, tipo_evento, tipo_rede, ativo)

    stmt = select(Protocolo)
    if tipo_evento:
        stmt = stmt.where(Protocolo.tipo_evento == tipo_evento)
    if tipo_rede:
        stmt = stmt.where(Protocolo.tipo_rede == tipo_rede)
    if ativo is not None:
        stmt = stmt.where(Protocolo.ativo == ativo)
    stmt = stmt.order_by(desc(Protocolo.data_criacao))

    result = await db.execute(stmt)
    protocolos = result.scalars().all()

    headers, rows = _build_rows(protocolos)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Protocolos"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left")

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=protocolos.xlsx",
        },
    )
